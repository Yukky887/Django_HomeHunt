import os
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from catalog.models import Home

class Command(BaseCommand):
    help = 'Import homes from Excel file'

    def handle(self, *args, **options):
        # Путь к Excel-файлу
        excel_file_path = os.path.join(os.getcwd(), 'all.xlsx')

        # Загрузка файла
        wb = load_workbook(excel_file_path)
        ws = wb.active

        # Пропуск первой строки (заголовки)
        rows = ws.iter_rows(min_row=2, values_only=True)

        # Счетчик созданных объектов
        created_count = 0

        for row in rows:
            try:
                # Разбор данных из Excel
                title = row[0]  # Название
                price = row[1]  # Цена
                url = row[2]  # URL
                description = row[3]  # Описание
                views = row[4] if row[4] else 0  # Просмотров (может быть пустым)
                seller = row[6]  # Продавец
                address = row[7] # Адрес (может быть пустым)

                # Создание объявления
                home = Home.objects.create(
                    title=title,
                    price=price,
                    original_address=url,
                    description=description,
                    views_count=views,
                    adrs=address,
                    landlord=None  # Не создаем Landlord
                )

                # Вывод информации о созданном объекте
                self.stdout.write(self.style.SUCCESS(f'Created home: {home.title}'))
                created_count += 1

            except Exception as e:
                self.stderr.write(self.style.ERROR(f'Error creating home: {row[0]} - {e}'))

        self.stdout.write(self.style.SUCCESS(f'Successfully imported {created_count} homes'))